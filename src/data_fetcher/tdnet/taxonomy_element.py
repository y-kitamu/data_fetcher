"""EDINETタクソノミのに設定されている勘定科目等の要素を一覧取得する"""

import warnings
from pathlib import Path

from loguru import logger
from openpyxl import load_workbook

from ..constants import PROJECT_ROOT
from .constants import document_types
from .constants.schema import TaxonomyElement

# 決算短信サマリー用の定数
target_sheets_jp = [
    "日本基準（通期・連結）",
    "日本基準（通期・非連結）",
    "日本基準（四半期・連結）",
    "日本基準（四半期・非連結）",
    "日本基準（一般２Ｑ・連結）",
    "日本基準（一般２Ｑ・非連結）",
    "日本基準（特定２Ｑ・連結）",
    "日本基準（特定２Ｑ・非連結）",
]
target_sheets_us = [
    "米国基準（通期・連結）",
    "米国基準（四半期・連結）",
    "米国基準（２Ｑ・連結）",
]
target_sheets_ifrs = [
    "IFRS（通期・連結）",
    "IFRS（四半期・連結）",
    "IFRS（一般２Ｑ・連結）",
    "IFRS（特定２Ｑ・連結）",
]
target_sheets_rvfc = [
    "業績予想の修正",
]
target_sheets_rvdf = [
    "配当予想の修正",
]

# 決算短信サマリー以外の報告書用の定数
edjp_excel_path = PROJECT_ROOT / "data/tdnet/1f_AccountList.xlsx"
edjp_target_sheet_names = [
    "一般商工業",
    "建設業",
    "銀行・信託業",
    "銀行・信託業（特定取引勘定設置銀行）",
    "建設保証業",
    "第一種金融商品取引業",
    "生命保険業",
    "損害保険業",
    "鉄道事業",
    "海運事業",
    "高速道路事業",
    "電気通信事業",
    "電気事業",
    "ガス事業",
    "資産流動化業",
    "投資運用業",
    "投資業",
    "特定金融業",
    "社会医療法人",
    "学校法人",
    "商品先物取引業",
    "リース事業",
    "投資信託受益証券",
]
ifrs_excel_path = PROJECT_ROOT / "data/tdnet/1g_IFRS_ElementList.xlsx"
ifrs_target_sheet_names = ["詳細ツリー"]


def collect_all_taxonomies() -> dict[str, list[TaxonomyElement]]:
    """すべての報告書のタクソノミ要素一覧を取得"""
    warnings.simplefilter("ignore")
    elements = collect_reports_taxonomies(edjp_excel_path, edjp_target_sheet_names)
    elements |= collect_reports_taxonomies(ifrs_excel_path, ifrs_target_sheet_names)
    summary_elems = collect_summary_taxonomies()
    elements["決算短信サマリー"] = summary_elems
    elements["予想修正報告"] = summary_elems
    warnings.resetwarnings()
    return elements


def collect_reports_taxonomies(
    excel_path: Path, target_sheets: list[str]
) -> dict[str, list[TaxonomyElement]]:
    """決算短信サマリー以外の報告書のタクソノミ要素一覧を取得"""

    # excelから要素を抽出
    wb = load_workbook(excel_path)

    current_document = None
    current_header = None
    elements = {}
    for sheet_name in target_sheets:
        for row in wb[sheet_name].iter_rows(values_only=True):
            if row[0] is not None and "科目一覧" in row[0]:
                # new document section
                doc_str = row[0].replace("科目一覧", "").strip()
                document_type = [
                    dtype
                    for dtype in document_types
                    if any([alias in doc_str for alias in dtype.aliases])
                ]
                if len(document_type) == 1:
                    current_document = document_type[0].name
                else:
                    current_document = "Unknown Document Type"

                if current_document not in elements:
                    elements[current_document] = []
                    logger.debug(f"add taxonomies of document : {current_document}")
            elif current_document is not None:
                # within document section
                if row[0] is not None:
                    if "科目分類" in row[0] or "標準ラベル（日本語）" in row[0]:
                        current_header = row
                        continue

                if current_header is not None:
                    jpn_label = row[current_header.index("冗長ラベル（日本語）")]
                    eng_label = row[current_header.index("冗長ラベル（英語）")]
                    namespace = row[current_header.index("名前空間プレフィックス")]
                    element_id = row[current_header.index("要素名")]
                    period_type = row[current_header.index("periodType")]
                    if (
                        jpn_label is not None
                        and eng_label is not None
                        and namespace is not None
                        and element_id is not None
                        and period_type is not None
                    ):
                        elem = TaxonomyElement(
                            japanese_label=jpn_label,
                            english_label=eng_label,
                            namespace=namespace,
                            element_id=element_id,
                            period_type=period_type,
                        )
                        if elem not in elements[current_document]:
                            elements[current_document].append(elem)

    wb.close()
    return elements


def collect_summary_taxonomies() -> list[TaxonomyElement]:
    """決算短信のタクソノミ要素一覧を取得"""

    excel_path = PROJECT_ROOT / "data/tdnet/項目リスト_事業会社.xlsx"
    wb = load_workbook(excel_path)

    target_sheets = (
        target_sheets_jp
        + target_sheets_us
        + target_sheets_ifrs
        + target_sheets_rvdf
        + target_sheets_rvfc
    )

    elements = []
    current_header = None
    for sheet_name in target_sheets:
        for row in wb[sheet_name].iter_rows(values_only=True):
            if row[0] is not None and "管理状況" in row[0]:
                current_header = row
            elif current_header is not None:
                jpn_label = row[current_header.index("冗長ラベル（日本語）")]
                eng_label = row[current_header.index("冗長ラベル（英語）")]
                namespace = row[current_header.index("名前空間プレフィックス")]
                element_id = row[current_header.index("要素名")]
                period_type = row[current_header.index("periodType")]
                if (
                    jpn_label is not None
                    and eng_label is not None
                    and namespace is not None
                    and element_id is not None
                    and period_type is not None
                ):
                    elem = TaxonomyElement(
                        japanese_label=jpn_label,
                        english_label=eng_label,
                        namespace=namespace,
                        element_id=element_id,
                        period_type=period_type,
                    )
                    if elem not in elements:
                        elements.append(elem)
    return elements
